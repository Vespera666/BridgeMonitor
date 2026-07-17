#!/usr/bin/env python3
"""
从 监测点1.xlsx 生成持久化文件：
  - monitor_storage.txt  (监测点 + 传感器类型列)
  - sensor_storage.txt   (7 个传感器类型)
"""

import os
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(BASE, "build", "监测点1.xlsx")
OUT = os.path.join(BASE, "data")

# ── 读取监测点 ──────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)
sh = wb["监测点"]

points = []  # [(名称, 类型, 编号), ...]
for r in range(2, sh.max_row + 1):
    name = str(sh.cell(row=r, column=1).value).strip()
    ptype = str(sh.cell(row=r, column=2).value).strip()
    pid = str(sh.cell(row=r, column=3).value).strip()
    if pid and pid != "None":
        points.append((name, ptype, pid))

print(f"读取监测点: {len(points)} 条")

# ── 传感器类型定义 ──────────────────────────────────────────────
SENSOR_DEFS = [
    ("索力监测传感器",   "SL",  "索力计",    "SL",  "通用"),
    ("挠度传感器",       "ND",  "挠度计",    "ND",  "通用"),
    ("振动监测传感器",   "ZD",  "振动计",    "ZD",  "通用"),
    ("支座位移传感器",   "ZY",  "位移计",    "ZY",  "通用"),
    ("伸缩缝监测传感器", "LY",  "伸缩缝计",  "LY",  "通用"),
    ("风速风向传感器",   "FS",  "风速风向仪","FS",  "通用"),
    ("温湿度监测传感器", "WSD", "温湿度计",  "WSD", "通用"),
]

# 监测点类型 → 传感器类型映射
ptype_to_sensor = {
    "索力":     "索力监测传感器",
    "挠度":     "挠度传感器",
    "振动":     "振动监测传感器",
    "支座位移": "支座位移传感器",
    "伸缩缝":   "伸缩缝监测传感器",
    "风速风向": "风速风向传感器",
    "温湿度":   "温湿度监测传感器",
}

# ── 1. monitor_storage.txt ─────────────────────────────────────
monitor_lines = ["监测点编号,断面名称,安装日期,传感器类型"]
for name, ptype, pid in points:
    sensor_type = ptype_to_sensor.get(ptype, "未知类型")
    monitor_lines.append(f"{pid},{name},2023-01-01,{sensor_type}")

monitor_path = os.path.join(OUT, "monitor_storage.txt")
with open(monitor_path, "w", encoding="utf-8") as f:
    f.write("\n".join(monitor_lines) + "\n")
print(f"生成 monitor_storage.txt: {len(points)} 行")

# ── 2. sensor_storage.txt ──────────────────────────────────────
sensor_lines = ["设备名称,规格,型号,厂家,生产日期,采集频率(秒),传感器类型,绑定监测点"]
for sensor_type, code, device_name, model_prefix, vendor in SENSOR_DEFS:
    model = f"{model_prefix}-100"
    sensor_lines.append(
        f"{device_name},标准,{model},{vendor},2023-01-01,60,{sensor_type},未绑定"
    )

sensor_path = os.path.join(OUT, "sensor_storage.txt")
with open(sensor_path, "w", encoding="utf-8") as f:
    f.write("\n".join(sensor_lines) + "\n")
print(f"生成 sensor_storage.txt: {len(SENSOR_DEFS)} 个传感器")

# ── 复制到 build 目录（运行时读取路径）─────────────────────────
import shutil
build_out = os.path.join(BASE, "build", "Desktop_Qt_6_11_1_MinGW_64_bit-Debug")
for fn in ["monitor_storage.txt", "sensor_storage.txt"]:
    src = os.path.join(OUT, fn)
    dst = os.path.join(build_out, fn)
    shutil.copy2(src, dst)
    print(f"复制 {fn} → build/")

print("\n完成！重启程序后 ComboBox 将有 101 个监测点可选。")
